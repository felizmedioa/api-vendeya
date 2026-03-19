# ============================================================================
# service.py — Lógica de filleo
# ============================================================================
from pathlib import Path

from app.features.process_shipment.filleo.schemas import FilleoRequest
from app.shared.http_client import HttpClient
from openpyxl import load_workbook
from app.core.config import settings

# Ruta absoluta a la carpeta templates dentro de filleo
TEMPLATES_DIR = Path(__file__).resolve().parent / "templates"

def fill_book(archivo: str, datos: FilleoRequest) -> str:
    """
    Agrega una fila con los datos del envío al archivo Excel.
    Retorna la ruta del archivo guardado.
    """
    ruta = TEMPLATES_DIR / archivo
    wb = load_workbook(ruta)
    sheet = wb['Hoja1']

    # Escribir específicamente en la Fila 2 (A2 hasta M2)
    datos_fila = [int(datos.dni), int(datos.telefono), None, None, None, settings.AGENCIA_ORIGEN, datos.destino, settings.TIPO_PAQUETE, 0.1, 0.1, 0.1, 1, 1]
    
    for col_idx, valor in enumerate(datos_fila, start=1): # start=1 para columna A
        sheet.cell(row=2, column=col_idx, value=valor)
    wb.save(ruta)
    wb.close()

    return str(ruta)

async def fillear(client: HttpClient, ruta_archivo: str) -> dict:
    """
    Envía el archivo Excel como multipart/form-data.

    Args:
        client: Cliente HTTP con sesión activa.
        ruta_archivo: Ruta del archivo Excel a enviar.
    """
    client.verificar_sesion()
    headers = client.obtener_headers_ajax()
    # Remover Content-Type y Accept para que httpx maneje el multipart correctamente
    headers.pop("Content-Type", None)
    headers.pop("Accept", None)
    headers["Accept"] = "application/json"

    with open(ruta_archivo, "rb") as f:
        response = await client.client.post(
            "/import-excel",
            headers=headers,
            files={"file": (Path(ruta_archivo).name, f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        print(response.json())
    if(response.status_code != 200):
        raise Exception("Error al enviar el archivo")
    return 'Archivo enviado correctamente'
